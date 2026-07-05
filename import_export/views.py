from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render

from journal.models import TeacherAssignment
from journal.permissions import get_teacher_assignments

from .importers import GradebookImporter, StudentImporter, StructureImporter


def _is_admin_or_dean(user):
    return user.is_superuser or getattr(user, "role", None) in ("admin", "dean")


@login_required
def import_hub(request):
    """Головна сторінка імпорту — вибір типу."""
    if not _is_admin_or_dean(request.user):
        messages.error(request, "Доступ лише для деканату та адміністраторів.")
        return redirect("journal:dashboard")

    # Список журналів для імпорту оцінок
    assignments = TeacherAssignment.objects.select_related(
        "teacher", "discipline", "group", "semester"
    ).order_by("discipline__name", "group__name")

    return render(request, "import_export/import_hub.html", {"assignments": assignments})


@login_required
def import_students(request):
    """Імпорт списку студентів."""
    if not _is_admin_or_dean(request.user):
        messages.error(request, "Доступ заборонено.")
        return redirect("journal:dashboard")

    if request.method == "POST":
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            messages.error(request, "Завантажте файл.")
            return redirect("import_export:import_students")

        try:
            importer = StudentImporter(uploaded_file)
        except ValueError as e:
            messages.error(request, str(e))
            return redirect("import_export:import_students")

        if importer.errors:
            return render(request, "import_export/import_preview.html", {
                "import_type": "students",
                "errors": importer.errors,
                "preview_rows": [],
            })

        # Якщо натиснуто «Підтвердити»
        if request.POST.get("confirm") == "1":
            results = importer.execute()
            return render(request, "import_export/import_result.html", {
                "import_type": "students",
                "results": results,
                "preview_rows": importer.preview_rows,
            })

        # Показати превʼю
        return render(request, "import_export/import_preview.html", {
            "import_type": "students",
            "preview_rows": importer.preview_rows,
            "errors": importer.errors,
            "file_name": uploaded_file.name,
        })

    return render(request, "import_export/import_form.html", {
        "import_type": "students",
        "title": "Імпорт списку студентів",
        "description": "Завантажте CSV або Excel файл зі стовпцями: Прізвище, Імʼя, По батькові, Група, № зачітної книжки",
    })


@login_required
def import_structure(request):
    """Імпорт структури або дисциплін."""
    if not _is_admin_or_dean(request.user):
        messages.error(request, "Доступ заборонено.")
        return redirect("journal:dashboard")

    if request.method == "POST":
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            messages.error(request, "Завантажте файл.")
            return redirect("import_export:import_structure")

        try:
            importer = StructureImporter(uploaded_file)
        except ValueError as e:
            messages.error(request, str(e))
            return redirect("import_export:import_structure")

        if importer.errors:
            return render(request, "import_export/import_preview.html", {
                "import_type": "structure",
                "errors": importer.errors,
                "preview_rows": [],
            })

        if request.POST.get("confirm") == "1":
            results = importer.execute()
            return render(request, "import_export/import_result.html", {
                "import_type": "structure",
                "results": results,
                "sub_type": importer.import_type,
            })

        return render(request, "import_export/import_preview.html", {
            "import_type": "structure",
            "sub_type": importer.import_type,
            "preview_rows": importer.preview_rows,
            "errors": importer.errors,
            "file_name": uploaded_file.name,
        })

    return render(request, "import_export/import_form.html", {
        "import_type": "structure",
        "title": "Імпорт структури / дисциплін",
        "description": "Завантажте CSV/Excel. Для структури: Факультет, Код факультету, Спеціальність, Код спеціальності, Група, Курс. Для дисциплін: Назва дисципліни, Код дисципліни.",
    })


@login_required
def import_gradebook(request):
    """Імпорт журналу оцінок у конкретний TeacherAssignment."""
    user = request.user
    if not (_is_admin_or_dean(user) or getattr(user, "role", None) == "teacher"):
        messages.error(request, "Доступ заборонено.")
        return redirect("journal:dashboard")

    # Доступні журнали
    if _is_admin_or_dean(user):
        assignments = TeacherAssignment.objects.all()
    else:
        assignments = get_teacher_assignments(user)
    assignments = assignments.select_related("teacher", "discipline", "group", "semester")

    if request.method == "POST":
        uploaded_file = request.FILES.get("file")
        assignment_id = request.POST.get("assignment_id")

        if not uploaded_file or not assignment_id:
            messages.error(request, "Виберіть журнал та завантажте файл.")
            return redirect("import_export:import_gradebook")

        try:
            importer = GradebookImporter(uploaded_file, assignment_id=int(assignment_id))
        except ValueError as e:
            messages.error(request, str(e))
            return redirect("import_export:import_gradebook")

        if importer.errors:
            return render(request, "import_export/import_preview.html", {
                "import_type": "gradebook",
                "errors": importer.errors,
                "preview_rows": [],
                "assignments": assignments,
            })

        if request.POST.get("confirm") == "1":
            results = importer.execute()
            return render(request, "import_export/import_result.html", {
                "import_type": "gradebook",
                "results": results,
            })

        return render(request, "import_export/import_preview.html", {
            "import_type": "gradebook",
            "preview_rows": importer.preview_rows,
            "column_lessons": importer.column_lessons,
            "max_grades": importer.max_grades,
            "errors": importer.errors,
            "file_name": uploaded_file.name,
            "assignment_id": assignment_id,
        })

    return render(request, "import_export/import_form.html", {
        "import_type": "gradebook",
        "title": "Імпорт журналу оцінок",
        "description": "Завантажте CSV/Excel з сіткою оцінок. Перший рядок — заголовки (Лекція 1, Лаб 2...), другий — макс. оцінки, далі — студенти.",
        "assignments": assignments,
    })


from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl
from journal.permissions import get_assignment_or_403
from university.models import Student
from journal.models import Lesson, Grade, Attendance

@login_required
def export_gradebook_excel(request, assignment_id):
    """Експортує вказаний журнал в Excel файл."""
    assignment = get_assignment_or_403(request.user, assignment_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Журнал оцінок"

    # Дані
    all_lessons = list(assignment.lessons.order_by("lesson_type", "order_number"))
    students = list(
        Student.objects.filter(group=assignment.group, is_active=True)
        .select_related("user")
        .order_by("user__last_name")
    )

    # Оцінки та відсутності
    grades_qs = Grade.objects.filter(lesson__assignment=assignment, grade_type=Grade.GradeType.CURRENT)
    grade_map = {(g.student_id, g.lesson_id): g.value for g in grades_qs}

    absent_qs = Attendance.objects.filter(lesson__assignment=assignment, status=Attendance.Status.ABSENT)
    absent_set = {(a.student_id, a.lesson_id) for a in absent_qs}

    # Стилі
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    font_danger = Font(name="Calibri", size=11, bold=True, color="842029")

    fill_header = PatternFill(start_color="212529", end_color="212529", fill_type="solid")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    fill_max = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    fill_absent = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    thin_side = Side(border_style="thin", color="DEE2E6")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    # Назва журналу
    ws.append([])
    ws.append([f"Електронний журнал: {assignment.discipline}"])
    ws.append([f"Група: {assignment.group} ({assignment.semester})"])
    ws.append([f"Викладач: {assignment.teacher.get_full_name()}"])
    ws.append([])

    for row in range(2, 5):
        ws.cell(row=row, column=1).font = font_bold

    start_row = 6

    # Заголовки типів занять
    ws.cell(row=start_row, column=1, value="Студент").font = font_header
    ws.cell(row=start_row, column=1).fill = fill_header
    ws.cell(row=start_row, column=1).alignment = align_center
    ws.cell(row=start_row, column=1).border = border_all

    ws.cell(row=start_row+1, column=1, value="").fill = fill_header
    ws.cell(row=start_row+1, column=1).border = border_all

    # Об'єднуємо назву "Студент" на два рядки
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row+1, end_column=1)

    # Групуємо заняття
    type_order = [
        (Lesson.Type.LECTURE, "Лекції"),
        (Lesson.Type.PRACTICE, "Практичні"),
        (Lesson.Type.LAB, "Лабораторні"),
    ]

    current_col = 2
    for lt, label in type_order:
        lessons_group = [l for l in all_lessons if l.lesson_type == lt]
        if lessons_group:
            ws.cell(row=start_row, column=current_col, value=label).font = font_header
            ws.cell(row=start_row, column=current_col).fill = fill_header
            ws.cell(row=start_row, column=current_col).alignment = align_center
            ws.cell(row=start_row, column=current_col).border = border_all

            if len(lessons_group) > 1:
                ws.merge_cells(start_row=start_row, start_column=current_col, end_row=start_row, end_column=current_col + len(lessons_group) - 1)

                for c in range(current_col + 1, current_col + len(lessons_group)):
                    ws.cell(row=start_row, column=c).fill = fill_header
                    ws.cell(row=start_row, column=c).border = border_all

            for idx, lesson in enumerate(lessons_group):
                col = current_col + idx
                ws.cell(row=start_row+1, column=col, value=lesson.order_number).font = font_header
                ws.cell(row=start_row+1, column=col).fill = fill_header
                ws.cell(row=start_row+1, column=col).alignment = align_center
                ws.cell(row=start_row+1, column=col).border = border_all

            current_col += len(lessons_group)

    # Стовпець Суми
    ws.cell(row=start_row, column=current_col, value="Σ").font = font_header
    ws.cell(row=start_row, column=current_col).fill = fill_header
    ws.cell(row=start_row, column=current_col).alignment = align_center
    ws.cell(row=start_row, column=current_col).border = border_all

    ws.cell(row=start_row+1, column=current_col, value="").fill = fill_header
    ws.cell(row=start_row+1, column=current_col).border = border_all
    ws.merge_cells(start_row=start_row, start_column=current_col, end_row=start_row+1, end_column=current_col)

    # Стовпець ECTS
    ws.cell(row=start_row, column=current_col + 1, value="ECTS").font = font_header
    ws.cell(row=start_row, column=current_col + 1).fill = fill_header
    ws.cell(row=start_row, column=current_col + 1).alignment = align_center
    ws.cell(row=start_row, column=current_col + 1).border = border_all

    ws.cell(row=start_row+1, column=current_col + 1, value="").fill = fill_header
    ws.cell(row=start_row+1, column=current_col + 1).border = border_all
    ws.merge_cells(start_row=start_row, start_column=current_col + 1, end_row=start_row+1, end_column=current_col + 1)

    # Стовпець Національна оцінка
    ws.cell(row=start_row, column=current_col + 2, value="Оцінка").font = font_header
    ws.cell(row=start_row, column=current_col + 2).fill = fill_header
    ws.cell(row=start_row, column=current_col + 2).alignment = align_center
    ws.cell(row=start_row, column=current_col + 2).border = border_all

    ws.cell(row=start_row+1, column=current_col + 2, value="").fill = fill_header
    ws.cell(row=start_row+1, column=current_col + 2).border = border_all
    ws.merge_cells(start_row=start_row, start_column=current_col + 2, end_row=start_row+1, end_column=current_col + 2)

    # Рядок Макс. оцінок
    max_row_idx = start_row + 2
    ws.cell(row=max_row_idx, column=1, value="⚙ Макс. оцінка").font = font_bold
    ws.cell(row=max_row_idx, column=1).fill = fill_max
    ws.cell(row=max_row_idx, column=1).border = border_all

    for idx, lesson in enumerate(all_lessons):
        col = 2 + idx
        max_val = lesson.max_grade if lesson.max_grade is not None else "—"
        cell = ws.cell(row=max_row_idx, column=col, value=max_val)
        cell.font = font_bold
        cell.fill = fill_max
        cell.alignment = align_center
        cell.border = border_all

    ws.cell(row=max_row_idx, column=current_col, value="—").font = font_bold
    ws.cell(row=max_row_idx, column=current_col).fill = fill_max
    ws.cell(row=max_row_idx, column=current_col).alignment = align_center
    ws.cell(row=max_row_idx, column=current_col).border = border_all

    for c in (current_col + 1, current_col + 2):
        ws.cell(row=max_row_idx, column=c, value="—").font = font_bold
        ws.cell(row=max_row_idx, column=c).fill = fill_max
        ws.cell(row=max_row_idx, column=c).alignment = align_center
        ws.cell(row=max_row_idx, column=c).border = border_all

    # Рядки студентів
    for s_idx, student in enumerate(students):
        r = max_row_idx + 1 + s_idx
        student_name = f"{student.user.last_name} {student.user.first_name}"
        ws.cell(row=r, column=1, value=student_name).font = font_regular
        ws.cell(row=r, column=1).alignment = align_left
        ws.cell(row=r, column=1).border = border_all

        total = 0
        for l_idx, lesson in enumerate(all_lessons):
            col = 2 + l_idx
            is_absent = (student.id, lesson.id) in absent_set
            grade_val = grade_map.get((student.id, lesson.id))

            cell = ws.cell(row=r, column=col)
            cell.alignment = align_center
            cell.border = border_all

            if is_absent:
                cell.value = "н"
                cell.font = font_danger
                cell.fill = fill_absent
            elif grade_val is not None:
                cell.value = grade_val
                cell.font = font_regular
                total += grade_val
            else:
                cell.value = ""

        # Загальний бал
        t_cell = ws.cell(row=r, column=current_col, value=total)
        t_cell.font = font_bold
        t_cell.alignment = align_center
        t_cell.border = border_all

        # Розрахунок ECTS та Національної
        percentage = (total / total_max_grade * 100) if total_max_grade > 0 else 0
        from journal.views import get_ects_and_national
        if total_max_grade > 0:
            ects, national = get_ects_and_national(percentage)
        else:
            ects, national = get_ects_and_national(total)

        # ECTS
        ects_cell = ws.cell(row=r, column=current_col + 1, value=ects)
        ects_cell.font = font_bold
        ects_cell.alignment = align_center
        ects_cell.border = border_all

        # Національна
        nat_cell = ws.cell(row=r, column=current_col + 2, value=national)
        nat_cell.font = font_regular
        nat_cell.alignment = align_center
        nat_cell.border = border_all

    # Вирівнювання ширини першого стовпчика
    max_len = max(len(f"{s.user.last_name} {s.user.first_name}") for s in students) if students else 20
    ws.column_dimensions["A"].width = max(max_len + 3, 20)

    for c in range(2, current_col + 3):
        col_letter = get_column_letter(c)
        if c == current_col + 2:
            ws.column_dimensions[col_letter].width = 18
        else:
            ws.column_dimensions[col_letter].width = 8

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"gradebook_{assignment.group.name}_{assignment.discipline.code}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response
